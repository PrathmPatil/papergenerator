from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


OPTION_RE = re.compile(r"\((\d)\)\s*(.+?)(?=\(\d\)|$)", re.DOTALL)
QUESTION_RE = re.compile(r"^(\d{1,3})[.)\s]\s*(.+)", re.DOTALL)
SKIP_PATTERNS = [
    r"^_c_",
    r"ALL[I1]M",
    r"^GEOMETRY EXERCISE",
    r"^Mathematics",
    r"^\d{3}-+$",
    r"^-{5,}",
    r"^={5,}",
    r"^\d+$",
]


@dataclass(frozen=True)
class PdfQuestionDefaults:
    class_id: str
    subject_id: str
    topic_id: str = ""
    difficulty: str = "easy"
    marks: int = 1
    negative_marks: float = 0


@dataclass
class ParsedQuestion:
    number: str
    text: str
    options: Dict[str, str]
    answer: str
    topic_name: str
    page_index: int
    passage: str = ""
    image_name: str = ""


class PdfQuestionExcelService:
    template_names = {
        "mcq_text": "mcq_text_questions_upload_template.xlsx",
        "mcq_image": "mcq_image_questions_template.xlsx",
        "paragraph": "paragraph_questions_upload_template.xlsx",
        "image_subquestions": "mcq_image_bulk_upload_template.xlsx",
    }

    output_names = {
        "mcq_text": "{prefix}_McqText.xlsx",
        "mcq_image": "{prefix}_McqImage.xlsx",
        "paragraph": "{prefix}_ParagraphQuestions.xlsx",
        "image_subquestions": "{prefix}_ImageSubquestions.xlsx",
    }

    fallback_headers = {
        "mcq_text": [
            "classId",
            "subjectId",
            "topicId",
            "type",
            "difficulty",
            "marks",
            "negativeMarks",
            "text",
            "optionA",
            "optionB",
            "optionC",
            "optionD",
            "correctAnswer",
        ],
        "mcq_image": [
            "classId",
            "subjectId",
            "topicId",
            "type",
            "difficulty",
            "marks",
            "negativeMarks",
            "questionText",
            "questionImage",
            "optionAText",
            "optionAImage",
            "optionBText",
            "optionBImage",
            "optionCText",
            "optionCImage",
            "optionDText",
            "optionDImage",
            "correctAnswer",
        ],
        "paragraph": [
            "classId",
            "subjectId",
            "topicId",
            "difficulty",
            "question_type",
            "paragraph_group_id",
            "instruction_text",
            "paragraph",
            "sub_question_id",
            "sub_question_type",
            "sub_question_text",
            "option_A",
            "option_B",
            "option_C",
            "option_D",
            "correct_answer",
            "marks",
            "negative_marks",
        ],
        "image_subquestions": [
            "classId",
            "subjectId",
            "topicId",
            "type",
            "question_group_id",
            "questionImage",
            "instructionText",
            "subQuestionId",
            "subQuestionText",
            "optionAText",
            "optionBText",
            "optionCText",
            "optionDText",
            "correctAnswer",
            "marks",
            "negativeMarks",
            "difficulty",
        ],
    }

    def generate_excel_zip(self, pdf_paths: Sequence[Path], defaults: PdfQuestionDefaults) -> BytesIO:
        parsed_questions: List[Dict] = []
        image_files: Dict[str, bytes] = {}
        image_counter = 0

        for pdf_path in pdf_paths:
            questions, pdf_images, image_counter = self._parse_pdf_with_notebook_logic(
                pdf_path,
                image_counter,
            )
            parsed_questions.extend(questions)
            image_files.update(pdf_images)

        unique_questions = self._dedupe_question_dicts(parsed_questions)
        categorized = self._categorize_notebook_questions(unique_questions)

        prefix = self._safe_output_prefix(defaults.class_id, defaults.subject_id)
        output = BytesIO()

        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as package:
            for key in ["mcq_text", "mcq_image", "paragraph", "image_subquestions"]:
                workbook_bytes = self._build_template_workbook(
                    template_key=key,
                    rows=self._build_rows(key, categorized[key], defaults),
                )
                package.writestr(self.output_names[key].format(prefix=prefix), workbook_bytes)
            if image_files:
                package.writestr("images.zip", self._build_images_zip(image_files).getvalue())

        output.seek(0)
        return output

    def _parse_pdf_with_notebook_logic(
        self,
        pdf_path: Path,
        image_counter: int = 0,
    ) -> Tuple[List[Dict], Dict[str, bytes], int]:
        questions: List[Dict] = []
        image_files: Dict[str, bytes] = {}
        current_q = None

        def save_q(question):
            if question and question.get("text"):
                questions.append(question)

        with fitz.open(str(pdf_path)) as pdf_doc:
            for page_index in range(len(pdf_doc)):
                page = pdf_doc.load_page(page_index)
                blocks = sorted(page.get_text("dict")["blocks"], key=lambda block: block["bbox"][1])

                for block in blocks:
                    if block["type"] == 1:
                        x0, y0, x1, y1 = block["bbox"]
                        if (x1 - x0) < 35 or (y1 - y0) < 35:
                            continue

                        image_name = f"page_{page_index + 1}_img_{image_counter}.png"
                        pix = page.get_pixmap(
                            matrix=fitz.Matrix(2.5, 2.5),
                            clip=fitz.Rect(block["bbox"]),
                        )
                        image_files[image_name] = pix.tobytes("png")
                        image_counter += 1

                        if current_q is not None:
                            current_q["images"].append(image_name)
                        continue

                    if block["type"] != 0:
                        continue

                    raw = " ".join(
                        span["text"]
                        for line in block.get("lines", [])
                        for span in line.get("spans", [])
                    ).strip()
                    raw = self._clean_text(raw)

                    if self._is_noise(raw):
                        continue

                    number, rest = self._is_question_start(raw)
                    if number is not None:
                        embedded_options = self._extract_options(rest)
                        question_text = re.split(r"\(1\)", rest)[0].strip() if embedded_options else rest

                        save_q(current_q)
                        current_q = {
                            "num": number,
                            "text": question_text,
                            "options": embedded_options,
                            "images": [],
                            "topic_name": self._topic_from_filename(pdf_path.name),
                        }
                        continue

                    if current_q is not None:
                        new_options = self._extract_options(raw)
                        if new_options:
                            current_q["options"].extend(new_options)
                        elif re.match(r"^\(\d\)", raw):
                            current_q["options"].append(re.sub(r"^\(\d\)\s*", "", raw).strip())

                if page_index == len(pdf_doc) - 1:
                    save_q(current_q)

        for question in questions:
            seen = set()
            deduped = []
            for option in question["options"]:
                if option not in seen:
                    seen.add(option)
                    deduped.append(option)
            question["options"] = deduped[:4]

        return questions, image_files, image_counter

    def _extract_options(self, text: str) -> List[str]:
        matches = OPTION_RE.findall(text or "")
        return [match[1].strip() for match in matches[:4]]

    def _is_question_start(self, text: str) -> Tuple[str | None, str | None]:
        text = str(text or "").strip()
        match = QUESTION_RE.match(text)
        if not match:
            return None, None

        number = match.group(1)
        rest = match.group(2).strip()
        if re.match(r"^\d+$", rest):
            return None, None

        return number, rest

    def _is_noise(self, text: str) -> bool:
        value = str(text or "").strip()
        if not value:
            return True
        return any(re.search(pattern, value) for pattern in SKIP_PATTERNS)

    def _dedupe_question_dicts(self, questions: Sequence[Dict]) -> List[Dict]:
        unique = []
        seen = set()

        for question in questions:
            key = re.sub(r"[^a-z0-9]+", "", str(question.get("text", "")).lower())
            if not key or key in seen:
                continue
            seen.add(key)
            unique.append(question)

        return unique

    def _build_images_zip(self, image_files: Dict[str, bytes]) -> BytesIO:
        output = BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as images_zip:
            for image_name, image_bytes in image_files.items():
                images_zip.writestr(image_name, image_bytes)
        output.seek(0)
        return output

    def _parse_pdf(self, pdf_path: Path) -> List[ParsedQuestion]:
        topic_name = self._topic_from_filename(pdf_path.name)
        questions: List[ParsedQuestion] = []
        answer_key: Dict[str, str] = {}
        passage_buffer: List[str] = []
        active_passage = ""
        current: Optional[Dict] = None

        def flush_current() -> None:
            nonlocal current
            if not current:
                return

            text = self._clean_text(" ".join(current["question_lines"]))
            options = self._fill_options(current["options"])
            answer = answer_key.get(current["number"], current.get("answer") or "A")
            questions.append(
                ParsedQuestion(
                    number=current["number"],
                    text=text,
                    options=options,
                    answer=self._normalize_answer(answer),
                    topic_name=topic_name,
                    page_index=current["page_index"],
                    passage=current.get("passage", ""),
                    image_name=current.get("image_name", ""),
                )
            )
            current = None

        with fitz.open(str(pdf_path)) as document:
            for page_index, page in enumerate(document):
                page_images = self._page_image_names(page, page_index)
                lines = self._extract_page_lines(page)

                for line in lines:
                    cleaned = self._clean_text(line)
                    if not cleaned:
                        continue

                    answer_match = self._match_answer(cleaned)
                    if answer_match and current is None:
                        answer_key[answer_match[0]] = answer_match[1]
                        continue
                    if answer_match and current is not None and answer_match[0] in {"", current["number"]}:
                        current["answer"] = answer_match[1]
                        continue

                    question_match = QUESTION_PATTERN.match(cleaned)
                    option_match = OPTION_PATTERN.match(cleaned)

                    if question_match:
                        flush_current()
                        if passage_buffer and len(" ".join(passage_buffer)) >= 120:
                            active_passage = self._clean_text("\n".join(passage_buffer))
                        passage_buffer = []

                        question_number = question_match.group(1)
                        question_text = question_match.group(2)
                        current = {
                            "number": question_number,
                            "question_lines": [question_text],
                            "options": {},
                            "answer": "",
                            "topic_name": topic_name,
                            "page_index": page_index,
                            "passage": active_passage,
                            "image_name": page_images[0] if page_images else "",
                        }
                        continue

                    if current and option_match:
                        current["options"][option_match.group(1).upper()] = option_match.group(2)
                        continue

                    if current:
                        current["question_lines"].append(cleaned)
                    else:
                        passage_buffer.append(cleaned)

            flush_current()

        return [question for question in questions if question.text]

    def _extract_page_lines(self, page: fitz.Page) -> List[str]:
        raw_lines: List[tuple] = []
        text_page = page.get_text("dict")
        for block in text_page.get("blocks", []):
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                text = " ".join(span.get("text", "") for span in line.get("spans", []))
                if text.strip():
                    raw_lines.append((line.get("bbox", [0, 0, 0, 0])[1], text))

        return [line for _, line in sorted(raw_lines, key=lambda item: item[0])]

    def _page_image_names(self, page: fitz.Page, page_index: int) -> List[str]:
        names = []
        blocks = page.get_text("dict").get("blocks", [])
        image_index = 1
        for block in blocks:
            if block.get("type") == 1:
                names.append(f"page_{page_index + 1}_image_{image_index}.png")
                image_index += 1
        return names

    def _categorize_questions(self, questions: Sequence[ParsedQuestion]) -> Dict[str, List[ParsedQuestion]]:
        categorized = {
            "mcq_text": [],
            "mcq_image": [],
            "paragraph": [],
            "image_subquestions": [],
        }

        image_groups: Dict[str, List[ParsedQuestion]] = {}
        for question in questions:
            if question.image_name:
                image_groups.setdefault(f"{question.topic_name}|{question.page_index}|{question.image_name}", []).append(question)

        image_subquestion_ids = {
            id(question)
            for group in image_groups.values()
            if len(group) > 1
            for question in group
        }

        for question in questions:
            if question.passage:
                categorized["paragraph"].append(question)
            elif id(question) in image_subquestion_ids:
                categorized["image_subquestions"].append(question)
            elif question.image_name or self._looks_image_based(question.text):
                categorized["mcq_image"].append(question)
            else:
                categorized["mcq_text"].append(question)

        return categorized

    def _categorize_notebook_questions(self, questions: Sequence[Dict]) -> Dict[str, List[Dict]]:
        categorized = {
            "mcq_text": [],
            "mcq_image": [],
            "paragraph": [],
            "image_subquestions": [],
        }

        image_groups: Dict[str, List[int]] = {}
        for index, question in enumerate(questions):
            if question.get("images"):
                image_groups.setdefault(question["images"][0], []).append(index)

        shared_image_question_indexes = {
            index
            for indexes in image_groups.values()
            if len(indexes) >= 2
            for index in indexes
        }

        for index, question in enumerate(questions):
            if index in shared_image_question_indexes:
                categorized["image_subquestions"].append(question)
            elif question.get("images"):
                categorized["mcq_image"].append(question)
            else:
                categorized["mcq_text"].append(question)

        return categorized

    def _build_rows(
        self,
        template_key: str,
        questions: Sequence,
        defaults: PdfQuestionDefaults,
    ) -> List[List]:
        if questions and isinstance(questions[0], dict):
            return self._build_notebook_rows(template_key, questions, defaults)

        if template_key == "mcq_text":
            return [
                [
                    defaults.class_id,
                    defaults.subject_id,
                    question.topic_name,
                    "mcq_text",
                    defaults.difficulty,
                    defaults.marks,
                    defaults.negative_marks,
                    question.text,
                    question.options["A"],
                    question.options["B"],
                    question.options["C"],
                    question.options["D"],
                    question.answer,
                ]
                for question in questions
            ]

        if template_key == "mcq_image":
            return [
                [
                    defaults.class_id,
                    defaults.subject_id,
                    question.topic_name,
                    "mcq_image",
                    defaults.difficulty,
                    defaults.marks,
                    defaults.negative_marks,
                    question.text,
                    question.image_name,
                    question.options["A"],
                    "",
                    question.options["B"],
                    "",
                    question.options["C"],
                    "",
                    question.options["D"],
                    "",
                    question.answer,
                ]
                for question in questions
            ]

        if template_key == "paragraph":
            passage_ids: Dict[str, str] = {}
            rows: List[List] = []
            for question in questions:
                if question.passage not in passage_ids:
                    passage_ids[question.passage] = f"P{len(passage_ids) + 1}"
                rows.append(
                    [
                        defaults.class_id,
                        defaults.subject_id,
                        question.topic_name,
                        defaults.difficulty,
                        "paragraph",
                        passage_ids[question.passage],
                        "Read the passage and answer the following questions.",
                        question.passage,
                        question.number,
                        "mcq",
                        question.text,
                        question.options["A"],
                        question.options["B"],
                        question.options["C"],
                        question.options["D"],
                        question.answer,
                        defaults.marks,
                        defaults.negative_marks,
                    ]
                )
            return rows

        if template_key == "image_subquestions":
            image_group_ids: Dict[str, str] = {}
            rows: List[List] = []
            for question in questions:
                group_key = f"{question.topic_name}|{question.page_index}|{question.image_name}"
                if group_key not in image_group_ids:
                    image_group_ids[group_key] = f"IMG{len(image_group_ids) + 1}"
                rows.append(
                    [
                        defaults.class_id,
                        defaults.subject_id,
                        question.topic_name,
                        "image_subquestions",
                        image_group_ids[group_key],
                        question.image_name,
                        "Observe the image and answer the following questions.",
                        question.number,
                        question.text,
                        question.options["A"],
                        question.options["B"],
                        question.options["C"],
                        question.options["D"],
                        question.answer,
                        defaults.marks,
                        defaults.negative_marks,
                        defaults.difficulty,
                    ]
                )
            return rows

        return []

    def _build_notebook_rows(
        self,
        template_key: str,
        questions: Sequence[Dict],
        defaults: PdfQuestionDefaults,
    ) -> List[List]:
        def pad4(options):
            padded = list(options or []) + [""] * 4
            return padded[0], padded[1], padded[2], padded[3]

        if template_key == "mcq_text":
            rows = []
            for question in questions:
                option_a, option_b, option_c, option_d = pad4(question.get("options"))
                topic_id = defaults.topic_id or question.get("topic_name") or "General"
                rows.append(
                    [
                        defaults.class_id,
                        defaults.subject_id,
                        topic_id,
                        "mcq_text",
                        defaults.difficulty,
                        defaults.marks,
                        defaults.negative_marks,
                        question.get("text", ""),
                        option_a,
                        option_b,
                        option_c,
                        option_d,
                        "",
                    ]
                )
            return rows

        if template_key == "mcq_image":
            rows = []
            for question in questions:
                option_a, option_b, option_c, option_d = pad4(question.get("options"))
                image_name = question.get("images", [""])[0] if question.get("images") else ""
                topic_id = defaults.topic_id or question.get("topic_name") or "General"
                rows.append(
                    [
                        defaults.class_id,
                        defaults.subject_id,
                        topic_id,
                        "mcq_image",
                        defaults.difficulty,
                        defaults.marks,
                        defaults.negative_marks,
                        question.get("text", ""),
                        image_name,
                        option_a,
                        "",
                        option_b,
                        "",
                        option_c,
                        "",
                        option_d,
                        "",
                        "",
                    ]
                )
            return rows

        if template_key == "image_subquestions":
            rows = []
            for question in questions:
                option_a, option_b, option_c, option_d = pad4(question.get("options"))
                image_name = question.get("images", [""])[0] if question.get("images") else ""
                topic_id = defaults.topic_id or question.get("topic_name") or "General"
                rows.append(
                    [
                        defaults.class_id,
                        defaults.subject_id,
                        topic_id,
                        "mcq_image",
                        f"IMG-GRP-{image_name.replace('.png', '')}",
                        image_name,
                        "Look at the figure and answer the following questions.",
                        f"SQ{question.get('num', '')}",
                        question.get("text", ""),
                        option_a,
                        option_b,
                        option_c,
                        option_d,
                        "",
                        defaults.marks,
                        defaults.negative_marks,
                        defaults.difficulty,
                    ]
                )
            return rows

        if template_key == "paragraph":
            return []

        return []

    def _build_template_workbook(self, template_key: str, rows: List[List]) -> bytes:
        workbook = self._load_template(template_key)
        sheet = workbook.active
        self._clear_data_rows(sheet)

        for row_index, row in enumerate(rows, start=2):
            for column_index, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _load_template(self, template_key: str):
        template_path = self._find_template_path(self.template_names[template_key])
        if template_path and template_path.exists():
            return load_workbook(template_path)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Questions"
        fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        for column_index, header in enumerate(self.fallback_headers[template_key], start=1):
            cell = sheet.cell(row=1, column=column_index, value=header)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        return workbook

    def _find_template_path(self, template_name: str) -> Optional[Path]:
        current = Path(__file__).resolve()
        candidates = [
            current.parents[1] / "templates" / "sample_file" / template_name,
            current.parents[4] / "frontend" / "public" / "sample_file" / template_name,
            current.parents[3] / "frontend" / "public" / "sample_file" / template_name,
            Path.cwd().parent / "frontend" / "public" / "sample_file" / template_name,
            Path.cwd() / "frontend" / "public" / "sample_file" / template_name,
        ]
        return next((path for path in candidates if path.exists()), None)

    def _clear_data_rows(self, sheet) -> None:
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)

    def _dedupe_questions(self, questions: Sequence[ParsedQuestion]) -> List[ParsedQuestion]:
        seen = set()
        unique = []
        for question in questions:
            key = re.sub(r"[^a-z0-9]+", "", question.text.lower())
            if not key or key in seen:
                continue
            seen.add(key)
            unique.append(question)
        return unique

    def _fill_options(self, options: Dict[str, str]) -> Dict[str, str]:
        cleaned = {key: self._clean_text(value) for key, value in options.items()}
        for key in ["A", "B", "C", "D"]:
            if not cleaned.get(key):
                cleaned[key] = f"Option {key}"
        return cleaned

    def _match_answer(self, line: str) -> Optional[tuple]:
        for pattern in ANSWER_PATTERNS:
            match = pattern.match(line)
            if not match:
                continue
            groups = match.groups()
            if len(groups) == 2:
                return groups[0], self._normalize_answer(groups[1])
            return "", self._normalize_answer(groups[0])
        return None

    def _normalize_answer(self, value: str) -> str:
        answer = str(value or "A").strip().upper()[:1]
        return answer if answer in {"A", "B", "C", "D", "E"} else "A"

    def _looks_image_based(self, text: str) -> bool:
        lowered = text.lower()
        image_words = ["diagram", "figure", "image", "picture", "chart", "graph", "observe", "shown below"]
        return any(word in lowered for word in image_words)

    def _topic_from_filename(self, filename: str) -> str:
        stem = Path(filename).stem
        stem = re.sub(r"^\s*[\d\W_]+", "", stem)
        stem = re.sub(r"[_\-]+", " ", stem)
        stem = re.sub(r"\s+", " ", stem).strip()
        return stem or "General"

    def _safe_output_prefix(self, class_id: str, subject_id: str) -> str:
        raw = f"{class_id}_{subject_id}".strip("_") or "QuestionBank"
        return re.sub(r"[^A-Za-z0-9]+", "_", raw).strip("_") or "QuestionBank"

    def _clean_text(self, value: str) -> str:
        text = str(value or "").replace("\xa0", " ")
        text = text.replace("�", "")
        text = re.sub(r"[ \t\r\f\v]+", " ", text)
        text = re.sub(r"\s+([,.;:?!])", r"\1", text)
        return text.strip()
