#!/usr/bin/env python3
"""Generate corrected paragraph Excel template with correct columns and values."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Paragraph Questions"

# Define headers
headers = [
    "classId",           # e.g., "10"
    "subjectId",         # e.g., "science"
    "topicId",           # e.g., "physics_motion"
    "difficulty",        # e.g., "easy", "medium", "hard"
    "question_type",     # MUST be "paragraph"
    "paragraph_group_id", # e.g., "P1" - same for all rows of same paragraph
    "instruction_text",  # e.g., "Read the passage and answer..."
    "paragraph",         # The passage text (same for all rows in group)
    "sub_question_id",   # e.g., "1", "2", "3"
    "sub_question_type", # "mcq" or "mcq_text" or "true_false" or "short_answer"
    "sub_question_text", # The sub-question text
    "option_A",          # (for MCQ) Option A text
    "option_B",          # (for MCQ) Option B text
    "option_C",          # (for MCQ) Option C text
    "option_D",          # (for MCQ) Option D text
    "correct_answer",    # (for MCQ) "A", "B", "C", "D"; (for true_false) "true"/"false"; (for short_answer) answer text
    "marks",             # e.g., "2"
    "negative_marks",    # e.g., "0.5"
]

# Write headers (row 1)
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Sample data - Paragraph 1 with 3 MCQ sub-questions
paragraph_1 = "The water cycle is the continuous process of water moving from Earth's surface to the atmosphere and back. It consists of four main stages: evaporation, condensation, precipitation, and collection. Evaporation occurs when water from oceans, lakes, and rivers turns into water vapor due to heat from the sun. Condensation is when water vapor cools and forms clouds. Precipitation is when water falls back to Earth as rain, snow, or hail. Finally, collection occurs when water gathers in oceans and lakes."

sample_data = [
    # Group P1 - MCQ questions
    ["10", "science", "environmental_science", "easy", "paragraph", "P1", "Study the water cycle passage and answer the following questions.", paragraph_1, "1", "mcq", "What is the first stage of the water cycle?", "Condensation", "Evaporation", "Precipitation", "Collection", "B", "2", "0"],
    ["10", "science", "environmental_science", "easy", "paragraph", "P1", "Study the water cycle passage and answer the following questions.", paragraph_1, "2", "mcq", "Which process occurs when water vapor cools?", "Evaporation", "Condensation", "Collection", "Precipitation", "B", "2", "0"],
    ["10", "science", "environmental_science", "easy", "paragraph", "P1", "Study the water cycle passage and answer the following questions.", paragraph_1, "3", "true_false", "Precipitation is when water falls back to Earth.", "", "", "", "", "true", "1", "0"],
    
    # Group P2 - Different paragraph with MCQ
    ["10", "english", "reading", "medium", "paragraph", "P2", "Read the following excerpt and answer the questions.", "The Industrial Revolution marked a major turning point in human history. Beginning in Britain in the late 1700s, it transformed agrarian societies into industrial ones. Key inventions like the steam engine and textile machinery powered economic growth and urbanization.", "1", "mcq", "When did the Industrial Revolution begin?", "Early 1600s", "Late 1700s", "Early 1800s", "Mid 1800s", "B", "2", "0.5"],
    ["10", "english", "reading", "medium", "paragraph", "P2", "Read the following excerpt and answer the questions.", "The Industrial Revolution marked a major turning point in human history. Beginning in Britain in the late 1700s, it transformed agrarian societies into industrial ones. Key inventions like the steam engine and textile machinery powered economic growth and urbanization.", "2", "short_answer", "Name one key invention of the Industrial Revolution.", "", "", "", "", "steam engine", "3", "1"],
]

# Write sample data
for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Auto-adjust column widths
for col_idx, header in enumerate(headers, 1):
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(str(header)) + 2)

# Freeze header row
ws.freeze_panes = "A2"

# Save workbook
output_path = r"D:\Paper_New\papergenerator\frontend\public\sample_file\paragraph_questions_upload_template.xlsx"
wb.save(output_path)
print(f"✅ Template generated: {output_path}")
print(f"Columns: {', '.join(headers)}")
print(f"Sample rows: {len(sample_data)}")
